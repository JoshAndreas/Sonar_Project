import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from datetime import date

from writers.KwargsNamedArgument import KwargsNamedArgument


class ExcelOpenPyxl:
  #
  # WORKBOOK METHODS
  #

  workbook_file_name = ""

  def open_workbook(self, workbook_file_name):
    workbook = openpyxl.load_workbook(workbook_file_name)
    self.workbook_file_name = workbook_file_name
    return workbook

  def create_workbook(self, workbook_file_name):
    workbook = openpyxl.Workbook()
    self.workbook_file_name = workbook_file_name
    return workbook

  def close_workbook(self, workbook):
    if self.workbook_file_name is None:
      today = date.today()
      self.workbook_file_name = today.strftime("%Y%m%d%H%M%S")
    workbook.save(self.workbook_file_name)

  def add_worksheet(self, workbook, worksheet_name):
    worksheet = workbook.create_sheet(worksheet_name)
    worksheet.title = worksheet_name
    return worksheet

  def insert_worksheet(self, workbook, worksheet_name, worksheet_position):
    worksheet = workbook.create_sheet(worksheet_name, worksheet_position)
    worksheet.title = worksheet_name
    return worksheet

  def delete_worksheet(self, workbook, worksheet_name):
    sheet = workbook.get_sheet_by_name(worksheet_name)
    workbook.remove_sheet(sheet)
    return workbook

  def exist_worksheet(self, workbook, worksheet_name):
    pass

  def get_worksheet_by_name(self, workbook, worksheet_name):
    worksheet = workbook[worksheet_name]
    return worksheet

  # #
  # # WORKSHEET METHODS
  # #

  def autofit_columns(self, worksheet):
    for col in worksheet.columns:
      max_length = 0
      column = col[0].column_letter  # Get the column name
      for cell in col:
        try:  # Necessary to avoid error on empty cells
          if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
        except:
          pass
      adjusted_width = (max_length + 2) * 1.2
      worksheet.column_dimensions[column].width = adjusted_width

  def filter_columns(self, worksheet, column_range):
    filters = worksheet.auto_filter
    filters.ref = column_range

  def write(self, worksheet, row_index, col_index, data, style=None):
    worksheet.cell(row=row_index, column=col_index, value=data)
    col_letter = get_column_letter(col_index)
    if style:
      worksheet[f"{col_letter}{row_index}"].style = style
    return worksheet

  def write_number(self, worksheet, row_index, col_index, data, number_format=None, style=None):
    worksheet.cell(row=row_index, column=col_index, value=data)
    col_letter = get_column_letter(col_index)
    if number_format:
      worksheet[f"{col_letter}{row_index}"].number_format = number_format
    if style:
      worksheet[f"{col_letter}{row_index}"].style = style
    return worksheet

  def register_style(self, workbook, style_name, **kwargs):
    if style_name in workbook.named_styles:
      del workbook._named_styles[workbook.style_names.index(style_name)]

    style = NamedStyle(name=style_name)
    if KwargsNamedArgument.FONT in kwargs:
      font = kwargs[KwargsNamedArgument.FONT]
      style.font = Font(
        name=font.name,
        size=font.size,
        bold=font.bold,
        italic=font.italic,
        color=font.color,
      )

    if KwargsNamedArgument.PATTERN_FILL in kwargs:
      pattern_fill = kwargs[KwargsNamedArgument.PATTERN_FILL]
      style.fill = PatternFill(
        fill_type=pattern_fill.fill_type,
        start_color=pattern_fill.start_color,
        end_color=pattern_fill.end_color,
      )

    if KwargsNamedArgument.ALIGNMENT in kwargs:
      alignment = kwargs[KwargsNamedArgument.ALIGNMENT]
      style.alignment = Alignment(
        horizontal=alignment.horizontal,
        vertical=alignment.vertical,
        text_rotation=alignment.text_rotation,
        indent=alignment.indent,
      )

    workbook.add_named_style(style)
    return workbook
